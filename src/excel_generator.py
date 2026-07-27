import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class IndigoExcelReportGenerator:
    def __init__(self, df, stat_summary, nlp_summary):
        self.df = df
        self.stat_summary = stat_summary
        self.nlp_summary = nlp_summary

    def generate_report(self, output_path=None):
        if output_path is None:
            project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            reports_dir = os.path.join(project_root, 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            output_path = os.path.join(reports_dir, 'IndiGo_Executive_Analytics_Report_2026.xlsx')

        print(f"Generating Executive Excel 365 Report at: {output_path}...")
        wb = openpyxl.Workbook()
        
        # Define Styling Palette
        navy_header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="003366")
        kpi_num_font = Font(name="Calibri", size=14, bold=True, color="003366")
        sub_font = Font(name="Calibri", size=10, italic=True, color="555555")
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        card_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

        # -------------------------------------------------------------
        # SHEET 1: Executive KPI Dashboard
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Executive KPI Dashboard"
        ws1.views.sheetView[0].showGridLines = True
        
        ws1["A1"] = "IndiGo Airlines - Executive Data Analytics Dashboard 2026"
        ws1["A1"].font = title_font
        ws1["A2"] = "Automated Executive Performance & NPS Audit Report"
        ws1["A2"].font = sub_font
        
        # Calculate Key Metrics
        total_pax = len(self.df)
        total_rev = self.df['total_revenue_inr'].sum()
        otp_pct = (self.df['arrival_delay_min'] <= 15).mean() * 100
        promoters = (self.df['nps_category'] == 'Promoter').sum()
        detractors = (self.df['nps_category'] == 'Detractor').sum()
        nps = ((promoters - detractors) / total_pax) * 100

        # Create KPI Cards
        kpis = [
            ("Total Passengers", f"{total_pax:,}", "B4", "C5"),
            ("Total Revenue (INR)", f"₹{total_rev:,.2f}", "E4", "F5"),
            ("On-Time Performance (OTP)", f"{otp_pct:.1f}%", "H4", "I5"),
            ("Net Promoter Score (NPS)", f"{nps:+.1f}", "K4", "L5")
        ]

        for label, val, start_cell, end_cell in kpis:
            c1 = ws1[start_cell]
            c1.value = label
            c1.font = Font(name="Calibri", size=9, bold=True, color="555555")
            c1.alignment = Alignment(horizontal="center", vertical="center")
            c1.fill = card_fill
            
            # Put value in row below
            col_letter = start_cell[0]
            val_row = int(start_cell[1:]) + 1
            c2 = ws1[f"{col_letter}{val_row}"]
            c2.value = val
            c2.font = kpi_num_font
            c2.alignment = Alignment(horizontal="center", vertical="center")
            c2.fill = card_fill

        # Route Level Summary Table
        ws1["A7"] = "Route-Level Performance Summary"
        ws1["A7"].font = Font(name="Calibri", size=13, bold=True, color="003366")
        
        headers = ["Route", "Passenger Volume", "Total Revenue (INR)", "Avg Delay (min)", "OTP %", "NPS"]
        for col_num, h in enumerate(headers, 1):
            cell = ws1.cell(row=8, column=col_num)
            cell.value = h
            cell.font = header_font
            cell.fill = navy_header_fill
            cell.alignment = Alignment(horizontal="center")

        route_stats = self.df.groupby('route').agg(
            pax=('passenger_id', 'count'),
            rev=('total_revenue_inr', 'sum'),
            delay=('arrival_delay_min', 'mean'),
            otp=('arrival_delay_min', lambda x: (x <= 15).mean() * 100),
            nps=('nps_score', 'mean')
        ).reset_index()

        for row_idx, r in enumerate(route_stats.itertuples(), start=9):
            ws1.cell(row=row_idx, column=1, value=r.route).alignment = Alignment(horizontal="center")
            ws1.cell(row=row_idx, column=2, value=r.pax).number_format = "#,##0"
            ws1.cell(row=row_idx, column=3, value=r.rev).number_format = "₹#,##0.00"
            ws1.cell(row=row_idx, column=4, value=round(r.delay, 1)).number_format = "0.0"
            ws1.cell(row=row_idx, column=5, value=r.otp / 100).number_format = "0.0%"
            ws1.cell(row=row_idx, column=6, value=round(r.nps, 1)).number_format = "0.0"

        # -------------------------------------------------------------
        # SHEET 2: Statistical Modeling & SPSS Results
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Statistical Analysis (SPSS)")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "SPSS / Python Advanced Statistical Modeling Output"
        ws2["A1"].font = title_font
        
        # Write Univariate Stats Table
        ws2["A4"] = "1. Univariate Statistical Distributions"
        ws2["A4"].font = Font(name="Calibri", size=12, bold=True, color="003366")
        
        univ_df = self.stat_summary['univariate']
        for col_idx, col_name in enumerate(['Metric'] + list(univ_df.columns), 1):
            cell = ws2.cell(row=5, column=col_idx)
            cell.value = str(col_name).upper()
            cell.font = header_font
            cell.fill = navy_header_fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (idx_label, row_data) in enumerate(univ_df.iterrows(), start=6):
            ws2.cell(row=row_idx, column=1, value=str(idx_label)).alignment = Alignment(horizontal="left")
            for c_idx, val in enumerate(row_data, start=2):
                ws2.cell(row=row_idx, column=c_idx, value=str(val)).alignment = Alignment(horizontal="center")

        # -------------------------------------------------------------
        # SHEET 3: NLP Feedback & Text Analytics
        # -------------------------------------------------------------
        ws3 = wb.create_sheet(title="NLP Text Analytics")
        ws3.views.sheetView[0].showGridLines = True
        
        ws3["A1"] = "Customer Feedback Sentiment & Key Phrase Mining"
        ws3["A1"].font = title_font

        ws3["A4"] = "Sentiment Distribution"
        ws3["A4"].font = Font(name="Calibri", size=12, bold=True, color="003366")
        
        sent_dict = self.nlp_summary['sentiment_distribution_pct']
        ws3["A5"] = "Sentiment Category"
        ws3["B5"] = "Percentage (%)"
        ws3["A5"].font = header_font
        ws3["B5"].font = header_font
        ws3["A5"].fill = navy_header_fill
        ws3["B5"].fill = navy_header_fill

        for r_idx, (sent_cat, pct) in enumerate(sent_dict.items(), start=6):
            ws3.cell(row=r_idx, column=1, value=sent_cat).alignment = Alignment(horizontal="left")
            ws3.cell(row=r_idx, column=2, value=pct / 100).number_format = "0.00%"

        # Adjust Column Widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_path)
        print(f"Executive Excel Report created successfully: {output_path}")
        return output_path

if __name__ == '__main__':
    from src.data_loader import IndigoDataLoader
    from src.stat_analysis import StatisticalAnalyzer
    from src.nlp_analytics import NLPTextAnalytics

    df = IndigoDataLoader().load_from_db()
    stat_analyzer = StatisticalAnalyzer(df)
    nlp_analyzer = NLPTextAnalytics(df)

    stat_summary = {
        'univariate': stat_analyzer.univariate_analysis(),
        'bivariate': stat_analyzer.bivariate_analysis()
    }
    nlp_summary = nlp_analyzer.generate_nlp_summary()

    excel_gen = IndigoExcelReportGenerator(df, stat_summary, nlp_summary)
    excel_gen.generate_report()
